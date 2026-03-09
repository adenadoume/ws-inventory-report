"""
Excel upload API: parse sales/buys Excel and upsert to Supabase.
Column mapping: Κωδικός / code, Περιγραφή / description, Ποσότητα / qty, Αξία / value.
"""
import re
from fastapi import APIRouter, File, HTTPException, UploadFile

from app.supabase_client import get_supabase

BATCH = 500

router = APIRouter(prefix="/excel", tags=["excel"])


def extract_supplier(desc: str) -> str:
    """Extract supplier code from description: 'B172 BASKET...' → 'B172'."""
    m = re.match(r"^([A-Za-z]\d{1,3})\s", desc or "")
    return m.group(1).upper() if m else ""


def parse_sales_or_buys_sheet(contents: bytes) -> list[dict]:
    """Parse first sheet: columns 0=code, 1=description, 2=qty, 3=value (Greek or English headers)."""
    from openpyxl import load_workbook
    from io import BytesIO

    wb = load_workbook(BytesIO(contents), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=1, values_only=True))
    if not rows:
        return []

    # First row may be headers (Κωδικός, Περιγραφή, Ποσότητα, Αξία) or data
    start = 0
    if rows[0][0] is not None and isinstance(rows[0][0], str):
        s = str(rows[0][0]).strip().lower()
        if "κωδ" in s or "code" in s or "ωδ" in s:
            start = 1
    records = []
    for row in rows[start:]:
        code = str(row[0] or "").strip()
        if not code:
            continue
        description = str(row[1] or "").strip()
        qty = float(row[2] or 0)
        value = float(row[3] or 0)
        supplier = extract_supplier(description)
        records.append({
            "code": code,
            "description": description,
            "supplier": supplier,
            "qty_sold": qty,
            "value_sold": value,
        })
    return records


def parse_buys_rows(contents: bytes) -> list[dict]:
    """Same columns but for buys: qty_bought, value_bought."""
    raw = parse_sales_or_buys_sheet(contents)
    return [
        {
            "code": r["code"],
            "description": r["description"],
            "supplier": r["supplier"],
            "qty_bought": r["qty_sold"],
            "value_bought": r["value_sold"],
        }
        for r in raw
    ]


def parse_sales_rows(contents: bytes) -> list[dict]:
    """For sales: qty_sold, value_sold (keys already in parse_sales_or_buys_sheet)."""
    return parse_sales_or_buys_sheet(contents)


@router.post("/sales")
async def upload_sales_excel(file: UploadFile = File(...)):
    """Upload SALES Excel; clears ws_sales_2025 and inserts rows."""
    if not file.filename or not (file.filename.endswith(".xlsx") or file.filename.endswith(".xls")):
        raise HTTPException(400, "File must be .xlsx or .xls")
    contents = await file.read()
    try:
        records = parse_sales_rows(contents)
    except Exception as e:
        raise HTTPException(422, f"Parse error: {e}")
    if not records:
        return {"table": "ws_sales_2025", "inserted": 0, "message": "No rows"}

    sb = get_supabase()
    sb.table("ws_sales_2025").delete().neq("id", "00000000-0000-0000-0000-000000000000").execute()
    for i in range(0, len(records), BATCH):
        batch = records[i : i + BATCH]
        sb.table("ws_sales_2025").insert(batch).execute()
    return {"table": "ws_sales_2025", "inserted": len(records), "filename": file.filename}


@router.post("/buys")
async def upload_buys_excel(file: UploadFile = File(...)):
    """Upload BUYS Excel; clears ws_buys_2025 and inserts rows."""
    if not file.filename or not (file.filename.endswith(".xlsx") or file.filename.endswith(".xls")):
        raise HTTPException(400, "File must be .xlsx or .xls")
    contents = await file.read()
    try:
        records = parse_buys_rows(contents)
    except Exception as e:
        raise HTTPException(422, f"Parse error: {e}")
    if not records:
        return {"table": "ws_buys_2025", "inserted": 0, "message": "No rows"}

    sb = get_supabase()
    sb.table("ws_buys_2025").delete().neq("id", "00000000-0000-0000-0000-000000000000").execute()
    for i in range(0, len(records), BATCH):
        batch = records[i : i + BATCH]
        sb.table("ws_buys_2025").insert(batch).execute()
    return {"table": "ws_buys_2025", "inserted": len(records), "filename": file.filename}


@router.get("/health")
async def health():
    return {"status": "ok", "service": "excel-api"}
